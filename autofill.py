import sys
import subprocess
import importlib

# --- Auto-install required packages ---
required = ["selenium", "openpyxl", "requests", "bs4", "geopy"]
for package in required:
    try:
        importlib.import_module(package)
    except ImportError:
        print(f"📦 Installing {package} …")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", package])

import os
import time
import re
from urllib.parse import quote_plus, urlencode
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
import json
from copy import copy
from openpyxl.utils import get_column_letter
import logging
from selenium.webdriver.remote.remote_connection import LOGGER
from datetime import datetime, timedelta
from geopy.distance import geodesic
import csv
from io import StringIO
from datetime import datetime, date
from math import radians, cos, sin, asin, sqrt
import random

# Suppress Selenium logging
LOGGER.setLevel(logging.WARNING)
logging.getLogger('selenium').setLevel(logging.WARNING)
logging.getLogger('urllib3').setLevel(logging.WARNING)

_REDFIN_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.redfin.com/",
    "Origin": "https://www.redfin.com",
    "DNT": "1",
    "Connection": "keep-alive",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "Cache-Control": "no-cache"
}

# Create a session for persistent cookies
session = requests.Session()
session.headers.update(_REDFIN_HEADERS)


def get_coordinates_from_address(address: str) -> tuple:
    """Get lat/lng coordinates from address using multiple methods with caching"""
    try:
        clean_address = address.strip()

        # Method 1: Try using Nominatim (OpenStreetMap)
        print("🔄 Using Nominatim geocoding service...")
        try:
            nominatim_url = f"https://nominatim.openstreetmap.org/search?q={quote_plus(clean_address)}&format=json&limit=1&countrycodes=us"
            headers = {'User-Agent': 'Mozilla/5.0 (compatible; PropertyComps/1.0)'}

            response = requests.get(nominatim_url, headers=headers, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data:
                    lat = float(data[0]['lat'])
                    lng = float(data[0]['lon'])
                    print(f"✅ Found coordinates via Nominatim: {lat}, {lng}")
                    return lat, lng
        except Exception as e:
            print(f"⚠️ Nominatim geocoding failed: {e}")

        # Method 2: Try Census Geocoding Service
        print("🔄 Trying Census geocoding service...")
        try:
            census_url = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
            params = {
                'address': clean_address,
                'benchmark': 'Public_AR_Current',
                'format': 'json'
            }

            response = requests.get(census_url, params=params, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data.get('result', {}).get('addressMatches'):
                    coords = data['result']['addressMatches'][0]['coordinates']
                    lat = float(coords['y'])
                    lng = float(coords['x'])
                    print(f"✅ Found coordinates via Census: {lat}, {lng}")
                    return lat, lng
        except Exception as e:
            print(f"⚠️ Census geocoding failed: {e}")

        print(f"❌ Could not get coordinates for {address}")
        return None, None

    except Exception as e:
        print(f"❌ Error getting coordinates: {e}")
        return None, None


def calculate_distance_fallback(address1: str, address2: str) -> float:
    """Calculate distance between two addresses using geocoding with better error handling"""
    try:
        lat1, lng1 = get_coordinates_from_address(address1)
        lat2, lng2 = get_coordinates_from_address(address2)

        if lat1 and lng1 and lat2 and lng2:
            # Haversine formula
            lat1, lng1, lat2, lng2 = map(radians, [lat1, lng1, lat2, lng2])
            dlat = lat2 - lat1
            dlng = lng2 - lng1
            a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlng / 2) ** 2
            c = 2 * asin(sqrt(a))
            miles = 3959 * c  # Earth's radius in miles
            return round(miles, 2)
    except Exception as e:
        print(f"⚠️ Distance calculation failed: {e}")
        pass
    return 999.0  # Return large distance for failed calculations


def calculate_distance_from_coords(lat1, lng1, lat2, lng2):
    """Calculate distance between two coordinate pairs"""
    try:
        # Haversine formula
        lat1, lng1, lat2, lng2 = map(radians, [lat1, lng1, lat2, lng2])
        dlat = lat2 - lat1
        dlng = lng2 - lng1
        a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlng / 2) ** 2
        c = 2 * asin(sqrt(a))
        miles = 3959 * c  # Earth's radius in miles
        return round(miles, 2)
    except:
        return 999.0


def wait_and_find_elements(driver, selectors, timeout=10):
    """Try multiple selectors and wait for elements to load"""
    for selector in selectors:
        try:
            elements = WebDriverWait(driver, timeout).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, selector))
            )
            if elements:
                return elements
        except:
            continue
    return []


def extract_property_details(card_element):
    """Extract property details from a card element with improved parsing"""
    try:
        card_text = card_element.text.lower()
        card_html = card_element.get_attribute('outerHTML')

        data = {}

        # Try to extract address - multiple approaches
        address_patterns = [
            r'data-rf-test-name="[^"]*address[^"]*"[^>]*>([^<]+)',
            r'class="[^"]*address[^"]*"[^>]*>([^<]+)',
            r'<div[^>]*>[^<]*(\d+[^,]+(?:st|nd|rd|th|street|ave|avenue|dr|drive|rd|road|blvd|boulevard|ln|lane|ct|court|way|pl|place)[^<]*)</div>',
        ]

        for pattern in address_patterns:
            match = re.search(pattern, card_html, re.IGNORECASE)
            if match:
                data['address'] = match.group(1).strip()
                break

        # If no address found in HTML, try text content
        if not data.get('address'):
            lines = card_text.split('\n')
            for line in lines:
                if any(word in line for word in ['st', 'ave', 'dr', 'rd', 'blvd', 'ln', 'ct', 'way', 'pl']) and any(
                        char.isdigit() for char in line):
                    data['address'] = line.strip()
                    break

        # Extract price
        price_patterns = [
            r'\$([0-9,]+)',
            r'sold[^$]*\$([0-9,]+)',
            r'price[^$]*\$([0-9,]+)'
        ]

        for pattern in price_patterns:
            match = re.search(pattern, card_text, re.IGNORECASE)
            if match:
                price_str = match.group(1).replace(',', '')
                data['price'] = int(price_str)
                break

        # Extract beds
        beds_match = re.search(r'(\d+)\s*(?:bed|br)', card_text)
        data['beds'] = int(beds_match.group(1)) if beds_match else 0

        # Extract baths
        baths_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:bath|ba)', card_text)
        data['baths'] = float(baths_match.group(1)) if baths_match else 0

        # Extract sqft
        sqft_match = re.search(r'(\d{1,3}(?:,\d{3})*)\s*(?:sq\s*ft|sqft)', card_text)
        if sqft_match:
            sqft_str = sqft_match.group(1).replace(',', '')
            data['sqft'] = int(sqft_str)
        else:
            data['sqft'] = 0

        return data if data.get('address') else None

    except Exception as e:
        print(f"❌ Error extracting property details: {e}")
        return None


def search_redfin_selenium_improved(address: str, radius_miles: float = 1.0, days_back: int = 365) -> list:
    """Improved Selenium scraping with better element detection and waiting"""
    driver = None
    try:
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-logging")
        options.add_argument("--log-level=3")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--disable-web-security")
        options.add_argument("--allow-running-insecure-content")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        print("🌐 Starting improved browser scraping...")
        driver = webdriver.Chrome(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        # Set window size for better rendering
        driver.set_window_size(1920, 1080)

        # Navigate to Redfin and search
        print(f"🔍 Searching for: {address}")
        driver.get("https://www.redfin.com")

        # Wait for page to load
        time.sleep(random.uniform(2, 4))

        # Find and use search box with multiple selectors
        search_selectors = [
            "#search-box-input",
            "input[placeholder*='Search']",
            "input[name='search']",
            ".search-input"
        ]

        search_box = None
        for selector in search_selectors:
            try:
                search_box = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                )
                break
            except:
                continue

        if search_box:
            search_box.clear()
            search_box.send_keys(address)
            time.sleep(1)
            search_box.send_keys(Keys.RETURN)
            time.sleep(random.uniform(3, 5))
            print("✅ Address search completed")
        else:
            print("⚠️ Could not find search box")
            return []

        # Wait for search results to load
        time.sleep(random.uniform(3, 5))

        # Try to navigate to sold homes
        current_url = driver.current_url
        print(f"📍 Current URL: {current_url}")

        # Construct sold homes URL
        if '/sold' not in current_url:
            # Try different URL patterns for sold homes
            base_url = current_url.split('?')[0]
            sold_urls = [
                f"{base_url}/sold",
                f"{base_url}?sold={days_back}d",
                f"{current_url}&sold={days_back}d" if '?' in current_url else f"{current_url}?sold={days_back}d"
            ]

            for sold_url in sold_urls:
                try:
                    print(f"🔄 Trying sold URL: {sold_url}")
                    driver.get(sold_url)
                    time.sleep(random.uniform(4, 6))

                    # Check if we have property cards
                    test_selectors = [
                        ".SearchResultCard",
                        ".HomeCard",
                        "[data-rf-test-name*='card']",
                        ".PropertyCard"
                    ]

                    found_cards = False
                    for selector in test_selectors:
                        try:
                            cards = driver.find_elements(By.CSS_SELECTOR, selector)
                            if len(cards) > 0:
                                found_cards = True
                                break
                        except:
                            continue

                    if found_cards:
                        print(f"✅ Found property cards at: {sold_url}")
                        break

                except Exception as e:
                    print(f"⚠️ Failed to load {sold_url}: {e}")
                    continue

        # Wait for property cards to load with multiple selectors
        property_selectors = [
            ".SearchResultCard",
            ".HomeCard",
            "[data-rf-test-name*='card']",
            ".PropertyCard",
            "[class*='Card']",
            ".result-item",
            ".property-item"
        ]

        property_cards = []
        for selector in property_selectors:
            try:
                cards = WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, selector))
                )
                if cards:
                    property_cards = cards
                    print(f"✅ Found {len(cards)} property cards with selector: {selector}")
                    break
            except:
                continue

        if not property_cards:
            print("⚠️ No property cards found, trying scroll and wait...")
            # Try scrolling to load more content
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(2)

            # Try again with more generic selectors
            generic_selectors = [
                "[class*='card' i]",
                "[class*='result' i]",
                "[class*='property' i]",
                "[class*='home' i]"
            ]

            for selector in generic_selectors:
                try:
                    cards = driver.find_elements(By.CSS_SELECTOR, selector)
                    if len(cards) > 5:  # Reasonable number of cards
                        property_cards = cards
                        print(f"✅ Found {len(cards)} elements with generic selector: {selector}")
                        break
                except:
                    continue

        if not property_cards:
            print("❌ No property cards found after all attempts")
            return []

        # Process property cards
        sold_homes = []
        for i, card in enumerate(property_cards[:50]):  # Limit to first 50
            try:
                property_data = extract_property_details(card)
                if property_data and property_data.get('address') and property_data.get('price'):
                    # Calculate distance
                    distance = calculate_distance_fallback(address, property_data['address'])

                    if distance <= radius_miles and distance < 999:
                        property_data['distance'] = distance
                        sold_homes.append(property_data)
                        print(
                            f"✅ Property {len(sold_homes)}: {property_data['address'][:50]}... - ${property_data['price']:,} ({distance:.2f}mi)")
                    else:
                        print(f"⚠️ Property outside radius: {distance:.2f}mi")

            except Exception as e:
                print(f"❌ Error processing card {i}: {e}")
                continue

        print(f"✅ Found {len(sold_homes)} properties within {radius_miles} miles")
        return sold_homes

    except Exception as e:
        print(f"❌ Improved Selenium scraping failed: {e}")
        return []
    finally:
        if driver:
            driver.quit()


def try_redfin_api_alternative(lat: float, lng: float, radius_miles: float = 1.0, days_back: int = 365) -> list:
    """Try alternative Redfin API endpoints"""
    try:
        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_back)

        # Try different API endpoints
        api_endpoints = [
            "https://www.redfin.com/stingray/api/gis",
            "https://www.redfin.com/stingray/api/v1/search",
            "https://www.redfin.com/stingray/api/home/details"
        ]

        # Convert radius to approximate bounding box
        lat_delta = radius_miles * 0.0145  # Rough conversion
        lng_delta = radius_miles * 0.0145

        for endpoint in api_endpoints:
            try:
                print(f"📡 Trying API endpoint: {endpoint}")

                params = {
                    'al': '1',
                    'num_homes': '350',
                    'ord': 'redfin-recommended-asc',
                    'page_number': '1',
                    'sf': '1,2,3,5,6,7',
                    'sold_within_days': str(days_back),
                    'status': '9',  # Sold status
                    'uipt': '1,2,3,4,5,6',
                    'v': '8',
                    'lat_min': str(lat - lat_delta),
                    'lat_max': str(lat + lat_delta),
                    'lng_min': str(lng - lng_delta),
                    'lng_max': str(lng + lng_delta)
                }

                response = session.get(endpoint, params=params, timeout=15)

                if response.status_code == 200:
                    print(f"✅ Got response from {endpoint}")

                    # Try to parse as JSON
                    try:
                        data = response.json()
                        if isinstance(data, dict) and 'payload' in data:
                            homes = data['payload'].get('homes', [])
                            if homes:
                                print(f"✅ Found {len(homes)} homes in API response")
                                return homes
                    except json.JSONDecodeError:
                        pass

            except Exception as e:
                print(f"⚠️ API endpoint {endpoint} failed: {e}")
                continue

        return []

    except Exception as e:
        print(f"❌ Alternative API search failed: {e}")
        return []


def get_redfin_comps_enhanced(address: str,
                              radius_miles: float = 1,
                              sold_within_days: int = 365,
                              max_rows: int = 200) -> list[dict]:
    """Enhanced version with multiple fallback methods"""
    try:
        print(f"🔍 Getting coordinates for: {address}")
        lat, lng = get_coordinates_from_address(address)

        if not lat or not lng:
            print("❌ Could not get coordinates for address")
            return []

        print(f"📍 Coordinates: {lat}, {lng}")

        # Method 1: Try alternative API endpoints
        print("🔄 Trying alternative API endpoints...")
        api_results = try_redfin_api_alternative(lat, lng, radius_miles, sold_within_days)

        if api_results:
            print(f"✅ API returned {len(api_results)} results")
            # Process API results (implementation depends on API response structure)
            # This would need to be implemented based on actual API response format

        # Method 2: Use improved Selenium scraping
        print("🔄 Using improved Selenium scraping...")
        selenium_results = search_redfin_selenium_improved(address, radius_miles, sold_within_days)

        if selenium_results:
            comps = []
            for result in selenium_results[:max_rows]:
                try:
                    comp = {
                        "address": result.get('address', ''),
                        "soldDate": datetime.now().strftime('%Y-%m-%d'),
                        "price": result.get('price', 0),
                        "sqft": result.get('sqft', 0),
                        "ppsq": round(result.get('price', 0) / result.get('sqft', 1)) if result.get('sqft',
                                                                                                    0) > 0 else 0,
                        "beds": result.get('beds', 0),
                        "baths": result.get('baths', 0),
                        "lot": 0,
                        "dist": result.get('distance', 0),
                        "url": "",
                        "img": None
                    }
                    comps.append(comp)
                except Exception as e:
                    print(f"⚠️ Error processing result: {e}")
                    continue

            print(f"✅ Successfully processed {len(comps)} comparable sales")
            return comps

        print("❌ No comparable sales found")
        return []

    except Exception as e:
        print(f"❌ Error in enhanced get_redfin_comps: {e}")
        return []

def is_ad_element(element):
    """Check if an element is likely an advertisement"""
    try:
        # Check class names and attributes for ad indicators
        class_name = element.get_attribute('class') or ''
        data_attrs = element.get_attribute('outerHTML') or ''

        ad_indicators = ['ad', 'advertisement', 'sponsored', 'promo', 'banner']
        return any(indicator in class_name.lower() or indicator in data_attrs.lower()
                   for indicator in ad_indicators)
    except:
        return False


def extract_data_from_card(card):
    """Extract property data from a card element"""
    try:
        data = {}

        # Try to extract address
        address_selectors = [
            "[data-rf-test-name*='address']",
            ".address, .property-address",
            "[class*='address' i]"
        ]

        for selector in address_selectors:
            try:
                address_elem = card.find_element(By.CSS_SELECTOR, selector)
                data['address'] = address_elem.text.strip()
                break
            except:
                continue

        # Try to extract price
        price_selectors = [
            "[data-rf-test-name*='price']",
            ".price, .sold-price",
            "[class*='price' i]"
        ]

        for selector in price_selectors:
            try:
                price_elem = card.find_element(By.CSS_SELECTOR, selector)
                price_text = price_elem.text.strip()
                # Extract numeric price
                price_str = ''.join(filter(str.isdigit, price_text))
                data['price'] = int(price_str) if price_str else 0
                break
            except:
                continue

        # Extract other details (beds, baths, sqft)
        details_text = card.text.lower()

        # Extract beds
        import re
        beds_match = re.search(r'(\d+)\s*bed', details_text)
        data['beds'] = int(beds_match.group(1)) if beds_match else 0

        # Extract baths
        baths_match = re.search(r'(\d+(?:\.\d+)?)\s*bath', details_text)
        data['baths'] = float(baths_match.group(1)) if baths_match else 0

        # Extract sqft
        sqft_match = re.search(r'(\d{1,3}(?:,\d{3})*)\s*sq\s*ft', details_text)
        if sqft_match:
            sqft_str = sqft_match.group(1).replace(',', '')
            data['sqft'] = int(sqft_str)
        else:
            data['sqft'] = 0

        return data if data.get('address') else None

    except Exception as e:
        print(f"❌ Error extracting card data: {e}")
        return None


def extract_json_from_html(html_source):
    """Extract JSON data from HTML source"""
    try:
        # Look for common patterns where Redfin embeds data
        import re

        # Pattern 1: window.__REDFIN_STATE__
        state_pattern = r'window\.__REDFIN_STATE__\s*=\s*({.+?});'
        match = re.search(state_pattern, html_source, re.DOTALL)
        if match:
            return json.loads(match.group(1))

        # Pattern 2: Other embedded JSON
        json_pattern = r'({[^{]*"homes"[^}]*(?:{[^}]*}[^}]*)*})'
        matches = re.finditer(json_pattern, html_source)
        for match in matches:
            try:
                return json.loads(match.group(1))
            except:
                continue

        return None
    except Exception as e:
        print(f"❌ Error extracting JSON from HTML: {e}")
        return None


def extract_homes_from_json(json_data, lat, lng, radius_miles):
    """Extract homes from JSON data"""
    try:
        homes = []

        # Try different JSON structures
        if isinstance(json_data, dict):
            # Look for homes in various keys
            for key in ['homes', 'listings', 'properties', 'results']:
                if key in json_data:
                    homes_data = json_data[key]
                    if isinstance(homes_data, list):
                        for home in homes_data:
                            processed_home = process_home_json(home)
                            if processed_home:
                                homes.append(processed_home)

        return homes[:50]  # Limit results

    except Exception as e:
        print(f"❌ Error extracting homes from JSON: {e}")
        return []


def process_home_json(home_data):
    """Process individual home JSON data"""
    try:
        if not isinstance(home_data, dict):
            return None

        # Extract basic info
        address = home_data.get('address', {})
        if isinstance(address, dict):
            full_address = f"{address.get('line', '')} {address.get('city', '')} {address.get('state', '')} {address.get('zip', '')}"
        else:
            full_address = str(address)

        price = home_data.get('price', {})
        if isinstance(price, dict):
            price_value = price.get('value', 0)
        else:
            price_value = price or 0

        return {
            'address': full_address.strip(),
            'price': price_value,
            'beds': home_data.get('beds', 0),
            'baths': home_data.get('baths', 0),
            'sqft': home_data.get('sqFt', 0) or home_data.get('squareFeet', 0),
        }

    except Exception as e:
        print(f"❌ Error processing home JSON: {e}")
        return None


def parse_csv_response(csv_text, lat, lng, radius_miles):
    """Parse CSV response from Redfin API"""
    try:
        if not csv_text or len(csv_text.strip()) < 50:
            print("⚠️ CSV response too short or empty")
            return []

        lines = csv_text.strip().split('\n')
        if len(lines) < 2:
            return []

        # Parse header
        headers = [h.strip().strip('"') for h in lines[0].split(',')]
        homes = []

        for line in lines[1:]:
            try:
                values = [v.strip().strip('"') for v in line.split(',')]
                if len(values) != len(headers):
                    continue

                home_dict = dict(zip(headers, values))

                # Extract coordinates for distance calculation
                home_lat = float(home_dict.get('LATITUDE', 0))
                home_lng = float(home_dict.get('LONGITUDE', 0))

                if home_lat and home_lng:
                    # Calculate distance
                    distance = calculate_distance_from_coords(lat, lng, home_lat, home_lng)

                    if distance <= radius_miles:
                        homes.append({
                            'raw_data': home_dict,
                            'distance': distance
                        })
            except Exception as e:
                print(f"⚠️ Error parsing CSV line: {e}")
                continue

        print(f"✅ Parsed {len(homes)} homes from enhanced CSV")
        return homes

    except Exception as e:
        print(f"❌ Error parsing CSV: {e}")
        return []

def parse_redfin_home_data(home_data, distance):
    """Parse home data into standardized format"""
    try:
        # Handle different data structures
        if isinstance(home_data, dict):
            address = home_data.get('address', '')
            price = home_data.get('price', 0)

            # Try to extract numeric price if it's a string
            if isinstance(price, str):
                price_str = ''.join(filter(str.isdigit, price))
                price = int(price_str) if price_str else 0

            sqft = home_data.get('sqft', 0) or home_data.get('squareFeet', 0)

            return {
                "address": address,
                "soldDate": datetime.now().strftime('%Y-%m-%d'),
                "price": int(price),
                "sqft": int(sqft) if sqft else 0,
                "ppsq": round(price / sqft) if sqft and sqft > 0 else 0,
                "beds": home_data.get('beds', 0),
                "baths": home_data.get('baths', 0),
                "lot": 0,
                "dist": distance,
                "url": "",
                "img": None
            }

        return None

    except Exception as e:
        print(f"❌ Error parsing home data: {e}")
        return None


def search_redfin_sold_homes_enhanced(lat: float, lng: float, radius_miles: float = 1.0, days_back: int = 365) -> list:
    """Search for sold homes using multiple API approaches"""
    try:
        # Try multiple Redfin API endpoints
        api_urls = [
            # Primary search API
            "https://www.redfin.com/stingray/api/gis",
            # Alternative search endpoint
            "https://www.redfin.com/stingray/api/v1/search/rentals",
            # Map data endpoint
            "https://www.redfin.com/stingray/api/gis-csv"
        ]

        # Calculate bounding box around the coordinates
        miles_to_deg = radius_miles * 0.0145

        for api_url in api_urls:
            try:
                print(f"📡 Trying API: {api_url}")

                if 'gis-csv' in api_url:
                    # CSV format request
                    params = {
                        'al': '1',
                        'market': 'cleveland',
                        'num_homes': '350',
                        'ord': 'redfin-recommended-asc',
                        'page_number': '1',
                        'sf': '1,2,3,5,6,7',
                        'sold_within_days': str(days_back),
                        'status': '9',
                        'uipt': '1,2,3,4,5,6',
                        'v': '8',
                        'region_id': '39035',  # Cleveland region
                        'region_type': '2'
                    }
                else:
                    # JSON format request
                    params = {
                        'al': '1',
                        'market': 'cleveland',
                        'num_homes': '350',
                        'ord': 'redfin-recommended-asc',
                        'page_number': '1',
                        'sf': '1,2,3,5,6,7',
                        'sold_within_days': str(days_back),
                        'status': '9',
                        'uipt': '1,2,3,4,5,6',
                        'v': '8'
                    }

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'application/json, text/csv, */*',
                    'Referer': 'https://www.redfin.com/'
                }

                response = session.get(api_url, params=params, headers=headers, timeout=15)

                if response.status_code == 200:
                    print(f"✅ Got response from {api_url}, processing...")

                    # Try JSON first
                    try:
                        data = response.json()
                        if 'payload' in data and 'homes' in data['payload']:
                            homes = []
                            for home in data['payload']['homes']:
                                # Calculate distance from search center
                                home_lat = home.get('latLong', {}).get('latitude')
                                home_lng = home.get('latLong', {}).get('longitude')

                                distance = 999
                                if home_lat and home_lng:
                                    distance = calculate_distance_from_coords(lat, lng, home_lat, home_lng)

                                # Only include homes within radius
                                if distance <= radius_miles:
                                    homes.append({
                                        'raw_data': home,
                                        'distance': distance
                                    })

                            if homes:
                                print(f"✅ Found {len(homes)} homes within {radius_miles} miles via JSON")
                                return homes

                    except (json.JSONDecodeError, KeyError):
                        # Try CSV format
                        try:
                            csv_homes = parse_csv_response(response.text, lat, lng, radius_miles)
                            if csv_homes:
                                print(f"✅ Found {len(csv_homes)} homes via CSV")
                                return csv_homes
                        except Exception as e:
                            print(f"⚠️ CSV parsing failed: {e}")

            except Exception as e:
                print(f"⚠️ API {api_url} failed: {e}")
                continue

        return []

    except Exception as e:
        print(f"❌ Enhanced search failed: {e}")
        return []


def search_redfin_sold_homes_selenium_enhanced(address: str, radius_miles: float = 1.0, days_back: int = 365) -> list:
    """Enhanced Selenium scraping with direct address search"""
    driver = None
    try:
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-logging")
        options.add_argument("--log-level=3")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        print("🌐 Starting address-based browser scraping...")
        driver = webdriver.Chrome(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        # Start with Redfin home page and search for the address directly
        print(f"🔍 Searching for address: {address}")
        driver.get("https://www.redfin.com")
        time.sleep(3)

        # Find and use the search box
        try:
            search_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "search-box-input"))
            )
            search_box.clear()
            search_box.send_keys(address)
            search_box.send_keys(Keys.RETURN)
            time.sleep(3)
            print("✅ Address search completed")
        except Exception as e:
            print(f"⚠️ Could not use search box: {e}")
            # Fallback to direct URL construction
            search_url = f"https://www.redfin.com/stingray/do/location-autocomplete?location={quote_plus(address)}&start=0&count=10&v=2"
            driver.get(search_url)
            time.sleep(3)

        # Navigate to sold homes in the area
        try:
            # Look for "Recently Sold" or similar link
            sold_links = driver.find_elements(By.XPATH,
                                              "//a[contains(text(), 'Recently Sold') or contains(text(), 'Sold') or contains(@href, 'sold')]")
            if sold_links:
                sold_links[0].click()
                time.sleep(5)
                print("✅ Navigated to sold homes")
            else:
                # Try modifying URL to show sold homes
                current_url = driver.current_url
                if '/sold' not in current_url:
                    if '?' in current_url:
                        sold_url = current_url + f"&sold={days_back}d"
                    else:
                        sold_url = current_url + f"/filter/include=sold-{days_back}d"
                    driver.get(sold_url)
                    time.sleep(5)
                    print("✅ Modified URL to show sold homes")
        except Exception as e:
            print(f"⚠️ Could not navigate to sold homes: {e}")

        # Wait for property cards to load
        try:
            WebDriverWait(driver, 15).until(
                lambda driver: len([card for card in driver.find_elements(By.CSS_SELECTOR,
                                                                          ".HomeCard, [data-rf-test-name='MapHomeCard'], .SearchResultCard")
                                    if not is_ad_element(card)]) > 0
            )
            print("✅ Property cards loaded")
        except:
            print("⚠️ Timeout waiting for property cards")

        # Extract properties with multiple selectors
        property_selectors = [
            "[data-rf-test-name*='HomeCard'], [data-rf-test-name*='home-card']",
            ".HomeCard, .SearchResultCard",
            "[data-rf-test-name*='listing'], [data-rf-test-name*='property']",
            ".listingCard, .propertyCard",
            "[class*='Card'][class*='home' i]"
        ]

        property_cards = []
        for selector in property_selectors:
            try:
                cards = driver.find_elements(By.CSS_SELECTOR, selector)
                if cards:
                    filtered_cards = [card for card in cards if not is_ad_element(card)]
                    if filtered_cards:
                        property_cards = filtered_cards
                        print(f"✅ Found {len(filtered_cards)} property cards with selector: {selector}")
                        break
            except:
                continue

        if not property_cards:
            print("⚠️ No property cards found, trying page source extraction...")
            # Try to extract from page source
            page_source = driver.page_source

            # Look for JSON data embedded in the page
            json_data = extract_json_from_html(page_source)
            if json_data:
                homes = extract_homes_from_json(json_data, None, None, radius_miles)
                if homes:
                    print(f"✅ Extracted {len(homes)} homes from page source")
                    return homes

            return []

        # Process cards and filter by distance
        sold_homes = []
        for i, card in enumerate(property_cards[:50]):
            try:
                home_data = extract_data_from_card(card)
                if home_data and home_data.get('address'):
                    # Skip if address contains obvious non-local indicators
                    home_address = home_data['address'].lower()
                    if any(state in home_address for state in
                           [' ca ', ' california', ' fl ', ' florida', ' tx ', ' texas', ' ny ',
                            ' new york']) and 'oh' not in home_address and 'ohio' not in home_address:
                        print(f"⚠️ Skipping non-local property: {home_data['address'][:40]}...")
                        continue

                    # Calculate distance to subject property
                    distance = calculate_distance_fallback(address, home_data['address'])

                    # Only include if within radius and distance calculation succeeded
                    if distance <= radius_miles and distance < 999:
                        home_data['distance'] = distance
                        sold_homes.append(home_data)
                        print(f"✅ Property {len(sold_homes)}: {home_data['address'][:40]}... - {distance:.2f}mi")
                    else:
                        print(f"⚠️ Property too far or invalid: {home_data['address'][:40]}... - {distance:.2f}mi")

            except Exception as e:
                print(f"❌ Error processing card {i + 1}: {e}")
                continue

        print(f"✅ Found {len(sold_homes)} properties within {radius_miles} miles")
        return sold_homes

    except Exception as e:
        print(f"❌ Enhanced Selenium scraping failed: {e}")
        return []
    finally:
        if driver:
            driver.quit()


def _bucket(comps, r_min, r_max, d_min, d_max):
    """Filter comps by distance and date ranges with error handling"""
    filtered = []
    for c in comps:
        try:
            # Safe date parsing
            sold_date_str = c.get("soldDate", "")
            if not sold_date_str:
                continue

            # Handle different date formats
            try:
                if len(sold_date_str) >= 10:
                    sold_date = date.fromisoformat(sold_date_str[:10])
                else:
                    # Try parsing other formats
                    sold_date = datetime.strptime(sold_date_str, '%m/%d/%Y').date()
            except:
                continue  # Skip if date can't be parsed

            days_old = (date.today() - sold_date).days
            distance = c.get("dist", 0)

            # Validate distance and date values
            if not isinstance(distance, (int, float)) or distance < 0:
                continue
            if not isinstance(days_old, int) or days_old < 0:
                continue

            in_distance_range = r_min < distance <= r_max
            in_date_range = d_min <= days_old < d_max

            if in_distance_range and in_date_range:
                filtered.append(c)
        except Exception as e:
            print(f"⚠️ Error filtering comp: {e}")
            continue

    return filtered


def log_comp_buckets(address: str, comps: list[dict]):
    """Pretty-print the four requested buckets to stdout with error handling."""
    # Filter out comps with missing essential data and debug why they're invalid
    valid_comps = []
    invalid_reasons = []

    for i, comp in enumerate(comps):
        try:
            # Debug each comp
            reasons = []
            if not comp.get('address'):
                reasons.append("no address")
            if not comp.get('soldDate'):
                reasons.append("no soldDate")
            if not isinstance(comp.get('price', 0), (int, float)) or comp.get('price', 0) <= 0:
                reasons.append(f"invalid price: {comp.get('price')}")
            if not isinstance(comp.get('dist', 0), (int, float)):
                reasons.append(f"invalid dist: {comp.get('dist')}")

            if reasons:
                invalid_reasons.append(f"Comp {i + 1}: {', '.join(reasons)}")
            else:
                valid_comps.append(comp)
        except Exception as e:
            invalid_reasons.append(f"Comp {i + 1}: error checking - {e}")

    # Print debug info about invalid comps
    if invalid_reasons:
        print(f"🚨 Invalid comps found:")
        for reason in invalid_reasons[:5]:  # Show first 5
            print(f"   {reason}")
        if len(invalid_reasons) > 5:
            print(f"   ... and {len(invalid_reasons) - 5} more")

    buckets = [
        ("🔹 ≤0.5 mi & ≤6 mo", _bucket(valid_comps, 0, 0.5, 0, 181)),
        ("🔹 ≤0.5 mi & 6-12 mo", _bucket(valid_comps, 0, 0.5, 181, 366)),
        ("🔸 0.5-1 mi & ≤6 mo", _bucket(valid_comps, 0.5, 1, 0, 181)),
        ("🔸 0.5-1 mi & 6-12 mo", _bucket(valid_comps, 0.5, 1, 181, 366)),
    ]

    print(f"🔍 Total comps available: {len(comps)} (valid: {len(valid_comps)})")
    for i, comp in enumerate(valid_comps):
        try:
            days_old = (date.today() - date.fromisoformat(comp["soldDate"][:10])).days
            price = comp.get('price', 0)
            dist = comp.get('dist', 0)
            address_short = comp.get('address', 'Unknown')[:30]
            print(f"   Comp {i + 1}: {address_short}... | {dist:.2f}mi | {days_old} days old | ${price:,}")
        except Exception as e:
            print(f"   Comp {i + 1}: Error displaying comp data: {e}")

    print(f"\n🔍 Bucket results:")
    for i, (title, bucket_comps) in enumerate(buckets, 1):
        print(f"   Bucket {i} ({title.split('(')[0].strip()}): {len(bucket_comps)} items")

    print("\n" + "═" * 65)
    print(f"🏠  COMPARABLE SALES AROUND: {address.upper()}")
    print("═" * 65)

    for title, rows in buckets:
        if not rows:
            continue
        # Safe sorting with None handling
        rows.sort(key=lambda x: (x.get("ppsq") is None or x.get("ppsq", 0) <= 0, -x.get("ppsq", 0)))
        print(f"\n{title}  ({len(rows)} found, sorted by $/sq ft ↓)")
        for i, c in enumerate(rows, 1):
            try:
                ppsq = f"${c.get('ppsq', 0):.0f}/sf" if c.get('ppsq', 0) > 0 else "n/a"
                price = c.get('price', 0)
                beds = c.get('beds', 0)
                baths = c.get('baths', 0)
                sqft = c.get('sqft', 0)
                dist = c.get('dist', 0)
                sold_date = c.get('soldDate', '')[:10] if c.get('soldDate') else 'unknown'
                url = c.get('url', '')
                img = c.get('img') or 'no-img'

                print(f"{i:>2}. {dist:.2f} mi | "
                      f"{sold_date} | "
                      f"{ppsq:<8} | "
                      f"${price:,} | "
                      f"{beds}bd/{baths}ba | "
                      f"{sqft:,} sf | "
                      f"{url} | "
                      f"{img}")
            except Exception as e:
                print(f"{i:>2}. Error displaying row: {e}")
    print("\n📋  End of comps\n" + "═" * 65 + "\n")


# Replace the original functions with enhanced versions
def get_redfin_comps(address: str, radius_miles: float = 1, sold_within_days: int = 365, max_rows: int = 200) -> list[
    dict]:
    return get_redfin_comps_enhanced(address, radius_miles, sold_within_days, max_rows)


def parse_json_response(json_text: str, lat: float, lng: float, radius_miles: float) -> list:
    """Parse JSON response from Redfin API"""
    try:
        # Handle JSONP format
        if json_text.startswith('{}&&'):
            json_text = json_text[4:]

        data = json.loads(json_text)
        sold_homes = []

        # Navigate through different possible JSON structures
        homes_data = None
        if 'payload' in data:
            if 'homes' in data['payload']:
                homes_data = data['payload']['homes']
            elif 'sections' in data['payload']:
                for section in data['payload']['sections']:
                    if 'rows' in section:
                        homes_data = section['rows']
                        break

        if homes_data:
            for home in homes_data:
                try:
                    if 'latLong' in home:
                        home_lat = home['latLong']['latitude']
                        home_lng = home['latLong']['longitude']

                        distance = geodesic((lat, lng), (home_lat, home_lng)).miles

                        if distance <= radius_miles:
                            sold_homes.append({
                                'lat': home_lat,
                                'lng': home_lng,
                                'distance': distance,
                                'raw_data': home
                            })

                except Exception as e:
                    continue

        print(f"✅ Parsed {len(sold_homes)} homes from JSON")
        return sold_homes

    except Exception as e:
        print(f"❌ Error parsing JSON: {e}")
        return []


def search_redfin_simple_requests(address: str, days_back: int = 365) -> list:
    """Simple requests-based approach to find sold homes"""
    try:
        print("🔍 Trying simple requests approach...")

        # Clean up the address for URL
        clean_address = address.replace(",", "").replace(" ", "-").lower()

        # Try different URL patterns that Redfin might use
        urls_to_try = [
            f"https://www.redfin.com/city/7158/OH/Columbus/filter/include=sold-1yr,property-type=house",
            f"https://www.redfin.com/zipcode/43224/filter/include=sold-1yr",
            "https://www.redfin.com/stingray/api/gis?al=1&region_id=17151&region_type=6&sold_within_days=365&status=9&uipt=1,2,3,4,5,6,7,8",
            f"https://www.redfin.com/city/7158/OH/Columbus/recently-sold"
        ]

        for url in urls_to_try:
            try:
                print(f"📡 Trying: {url[:60]}...")
                response = session.get(url, timeout=15)

                if response.status_code == 200:
                    content = response.text
                    print(f"✅ Got response, length: {len(content)}")

                    # Look for JSON data in the response
                    if content.startswith('{}&&'):
                        json_content = content[4:]
                    else:
                        json_content = content

                    # Try to find property data
                    if '"homes"' in content or '"properties"' in content or '"listings"' in content:
                        print("✅ Found property data in response")
                        # This would need parsing - for now just return empty
                        return []

                    # Look for CSV-like data
                    if 'ADDRESS' in content and 'PRICE' in content:
                        print("✅ Found CSV-like data")
                        return parse_csv_response(content, 0, 0, 10)  # Large radius since no coords

            except Exception as e:
                print(f"⚠️ URL failed: {e}")
                continue

        return []

    except Exception as e:
        print(f"❌ Simple requests failed: {e}")
        return []



def search_redfin_url(address):
    """Search for a Redfin URL using DuckDuckGo"""
    print(f"🔍 Searching for Redfin listing: {address}")
    return try_duckduckgo_search(address)


def try_duckduckgo_search(address):
    """Try DuckDuckGo search for Redfin listing"""
    print(f"🦆 Searching with DuckDuckGo...")

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")

    try:
        service = ChromeService()
        driver = webdriver.Chrome(service=service, options=options)

        # Try multiple search query formats for better results
        queries = [
            f'{address} site:redfin.com',
            f'"{address}" site:redfin.com',
            f'{address.replace(",", "")} site:redfin.com',
            f'{address} redfin listing',
        ]

        for query in queries:
            try:
                ddg_url = f"https://duckduckgo.com/?q={quote_plus(query)}"
                print(f"🔍 Trying query: {query}")

                driver.get(ddg_url)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='redfin.com']"))
                )

                # Look for Redfin links in search results
                links = driver.find_elements(By.CSS_SELECTOR, "a[href*='redfin.com']")

                for link in links:
                    href = link.get_attribute("href")
                    if href and "redfin.com" in href and "/home/" in href:
                        print(f"✅ Found Redfin URL: {href}")
                        return href

                # If no results, try next query
                print(f"❌ No results for query: {query}")

            except Exception as e:
                print(f"⚠️ Error with query '{query}': {e}")
                continue

        print("❌ Could not find Redfin listing using DuckDuckGo")
        return None

    except Exception as e:
        print(f"⚠️ DuckDuckGo search failed: {e}")
        return None
    finally:
        try:
            driver.quit()
        except:
            pass


def is_valid_redfin_url(url):
    """Check if a URL is a valid Redfin property URL"""
    if not url or not isinstance(url, str):
        return False
    return url.startswith("http") and "redfin.com" in url and "/home/" in url


def get_redfin_data(url):
    print(f"🌐 Scraping Redfin data: {url}")
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")

    data = {}
    try:
        service = ChromeService()
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(url)

        wait = WebDriverWait(driver, 15)

        # --- Price ---
        try:
            price_el = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "[data-rf-test-id='abp-price']")))
            price_text = price_el.text.strip()
            match = re.search(r"\$([\d,]+)", price_text)
            if match:
                price_numeric = int(match.group(1).replace(",", ""))
                data["asking price (PP)"] = price_numeric
                print(f"💰 Price: ${price_numeric:,}")
            else:
                print("⚠️ Could not extract numeric price")
        except Exception as e:
            print(f"⚠️ Price not found: {e}")

        # --- Beds / Baths / SqFt / Garage - Enhanced extraction ---
        beds = None
        baths = None
        sqft = None
        garage = None

        # Try to extract from JSON data first (most reliable)
        try:
            json_match = re.search(r'"beds"\s*:\s*(\d+)', driver.page_source)
            if json_match:
                beds = json_match.group(1)
                print(f"🛏️ Beds (from JSON): {beds}")

            json_match = re.search(r'"baths"\s*:\s*([\d.]+)', driver.page_source)
            if json_match:
                baths = json_match.group(1)
                print(f"🛁 Baths (from JSON): {baths}")

            json_match = re.search(r'"sqFt"\s*:\s*(\d+)', driver.page_source)
            if json_match:
                sqft = int(json_match.group(1))
                print(f"📏 SqFt (from JSON): {sqft}")

            # Try to find garage in JSON
            garage_patterns = [
                r'"garage"\s*:\s*(\d+)',
                r'"garageSpaces"\s*:\s*(\d+)',
                r'"parkingSpaces"\s*:\s*(\d+)'
            ]
            for pattern in garage_patterns:
                garage_match = re.search(pattern, driver.page_source)
                if garage_match:
                    garage = garage_match.group(1)
                    print(f"🚗 Garage (from JSON): {garage}")
                    break

        except Exception as e:
            print(f"⚠️ JSON extraction failed: {e}")

        # Fallback to HTML parsing if JSON didn't work
        if not all([beds, baths, sqft]):
            try:
                # Enhanced statsValue parsing - get all stat values
                facts_block = driver.find_elements(By.CSS_SELECTOR, ".statsValue")
                clean_values = [v.text.strip() for v in facts_block if v.text.strip() and "$" not in v.text]
                print(f"📊 Found stats values: {clean_values}")

                # IMPROVED LOGIC: Try to identify beds/baths/sqft more reliably
                if len(clean_values) >= 2:
                    # Method 1: Use position-based logic with validation
                    potential_beds = clean_values[0] if len(clean_values) > 0 else None
                    potential_baths = clean_values[1] if len(clean_values) > 1 else None
                    potential_sqft = None

                    # Find the largest numeric value as likely sqft
                    for val in clean_values:
                        val_clean = re.sub(r"[^\d]", "", val)
                        if val_clean.isdigit() and int(val_clean) > 500:  # Reasonable sqft minimum
                            potential_sqft = int(val_clean)
                            break

                    # Validate and assign beds
                    if not beds and potential_beds and potential_beds.isdigit():
                        beds_num = int(potential_beds)
                        if 1 <= beds_num <= 10:  # Reasonable bed range
                            beds = potential_beds
                            print(f"🛏️ Beds (from statsValue position): {beds}")

                    # Validate and assign baths - IMPROVED LOGIC
                    if not baths and potential_baths:
                        # Handle both integer and decimal bath counts
                        if re.match(r'^\d+$', potential_baths):  # Integer like "2"
                            baths_num = int(potential_baths)
                            if 1 <= baths_num <= 10:  # Reasonable bath range
                                baths = potential_baths
                                print(f"🛁 Baths (from statsValue position): {baths}")
                        elif re.match(r'^\d+\.\d+$', potential_baths):  # Decimal like "2.5"
                            baths_num = float(potential_baths)
                            if 0.5 <= baths_num <= 10:  # Reasonable bath range
                                baths = potential_baths
                                print(f"🛁 Baths (from statsValue position): {baths}")

                    # Assign sqft
                    if not sqft and potential_sqft:
                        sqft = potential_sqft
                        print(f"📏 SqFt (from statsValue position): {sqft}")

                # Method 2: Try to find missing values with enhanced selectors
                if not beds:
                    bed_selectors = [
                        "[data-rf-test-id='abp-beds']",
                        ".beds .statsValue",
                        "[class*='bed']",
                        "span:contains('bed')",
                        "div:contains('bed')"
                    ]
                    for selector in bed_selectors:
                        try:
                            bed_el = driver.find_element(By.CSS_SELECTOR, selector)
                            bed_text = bed_el.text.strip()
                            bed_match = re.search(r'(\d+)', bed_text)
                            if bed_match and 1 <= int(bed_match.group(1)) <= 10:
                                beds = bed_match.group(1)
                                print(f"🛏️ Beds (from enhanced selector): {beds}")
                                break
                        except:
                            continue

                if not baths:
                    # ENHANCED BATH EXTRACTION with multiple strategies
                    bath_selectors = [
                        "[data-rf-test-id='abp-baths']",
                        ".baths .statsValue",
                        "[class*='bath']",
                        "span:contains('bath')",
                        "div:contains('bath')"
                    ]

                    for selector in bath_selectors:
                        try:
                            bath_el = driver.find_element(By.CSS_SELECTOR, selector)
                            bath_text = bath_el.text.strip()
                            # Look for patterns like "2 bath", "2.5 baths", "2 full baths"
                            bath_patterns = [
                                r'(\d+\.?\d*)\s*(?:full\s*)?baths?',
                                r'(\d+\.?\d*)\s*ba(?:th)?',
                                r'(\d+\.?\d*)'
                            ]

                            for pattern in bath_patterns:
                                bath_match = re.search(pattern, bath_text.lower())
                                if bath_match:
                                    bath_val = bath_match.group(1)
                                    try:
                                        bath_num = float(bath_val)
                                        if 0.5 <= bath_num <= 10:
                                            # Format properly (remove .0 for whole numbers)
                                            if bath_num == int(bath_num):
                                                baths = str(int(bath_num))
                                            else:
                                                baths = str(bath_num)
                                            print(f"🛁 Baths (from enhanced selector): {baths}")
                                            break
                                    except ValueError:
                                        continue
                            if baths:
                                break
                        except:
                            continue

                    # Additional strategy: Look for bath info in page text
                    if not baths:
                        try:
                            soup = BeautifulSoup(driver.page_source, "html.parser")
                            page_text = soup.get_text().lower()

                            # Look for patterns in the full page text
                            bath_text_patterns = [
                                r'(\d+\.?\d*)\s*(?:full\s*)?baths?',
                                r'(\d+\.?\d*)\s*ba(?:th)?',
                                r'baths?\s*:\s*(\d+\.?\d*)',
                                r'bath\s*count\s*:\s*(\d+\.?\d*)'
                            ]

                            for pattern in bath_text_patterns:
                                matches = re.findall(pattern, page_text)
                                for match in matches:
                                    try:
                                        bath_num = float(match)
                                        if 0.5 <= bath_num <= 10:
                                            # Format properly
                                            if bath_num == int(bath_num):
                                                baths = str(int(bath_num))
                                            else:
                                                baths = str(bath_num)
                                            print(f"🛁 Baths (from page text): {baths}")
                                            break
                                    except ValueError:
                                        continue
                                if baths:
                                    break
                        except Exception as e:
                            print(f"⚠️ Error in page text bath extraction: {e}")

                if not sqft:
                    sqft_selectors = [
                        "[data-rf-test-id='abp-sqFt']",
                        ".sqft .statsValue",
                        "[class*='sqft']",
                        "[class*='SqFt']"
                    ]
                    for selector in sqft_selectors:
                        try:
                            sqft_el = driver.find_element(By.CSS_SELECTOR, selector)
                            sqft_text = sqft_el.text.strip()
                            sqft_clean = re.sub(r"[^\d]", "", sqft_text)
                            if sqft_clean.isdigit() and int(sqft_clean) > 100:
                                sqft = int(sqft_clean)
                                print(f"📏 SqFt (from selector): {sqft}")
                                break
                        except:
                            continue

            except Exception as e:
                print(f"⚠️ Error extracting Beds/Baths/SqFt: {e}")

        # Enhanced garage extraction from multiple sources
        if not garage:
            try:
                # Look for garage in property features/details
                soup = BeautifulSoup(driver.page_source, "html.parser")
                page_text = soup.get_text().lower()

                # Search for garage patterns in text
                garage_patterns = [
                    r'(\d+)\s*car\s*garage',
                    r'garage\s*:\s*(\d+)',
                    r'(\d+)\s*garage',
                    r'parking\s*spaces?\s*:\s*(\d+)',
                    r'garage\s*spaces?\s*:\s*(\d+)'
                ]

                for pattern in garage_patterns:
                    match = re.search(pattern, page_text)
                    if match:
                        garage = match.group(1)
                        print(f"🚗 Garage (from text pattern): {garage}")
                        break

                # Also look in structured data sections
                if not garage:
                    detail_sections = soup.find_all(['div', 'span', 'li'],
                                                    string=re.compile(r'garage|parking', re.IGNORECASE))
                    for section in detail_sections:
                        parent_text = section.get_text() if section.parent else ""
                        match = re.search(r'(\d+)', parent_text)
                        if match and 1 <= int(match.group(1)) <= 10:
                            garage = match.group(1)
                            print(f"🚗 Garage (from detail section): {garage}")
                            break

            except Exception as e:
                print(f"⚠️ Error extracting garage: {e}")

        # Store the extracted values
        if sqft:
            data["sqft"] = sqft

        # Build property type string with beds/baths/garage - FIXED FORMATTING
        property_type_parts = ["SFR"]

        if beds:
            property_type_parts.append(beds)
        else:
            property_type_parts.append("?")

        if baths:
            # Format baths to remove unnecessary decimal places
            if '.' in str(baths) and str(baths).endswith('.0'):
                baths_formatted = str(int(float(baths)))
            else:
                baths_formatted = str(baths)
            property_type_parts.append(baths_formatted)
        else:
            property_type_parts.append("?")

        if garage:
            property_type_parts.append(garage)
        else:
            property_type_parts.append("0")  # Default to 0 if no garage found

        # FIXED: Use the exact label from the Excel sheet
        property_type_str = f"{property_type_parts[0]} {property_type_parts[1]}/{property_type_parts[2]}/{property_type_parts[3]}"
        data["property type + bd/bt/garage (example: SFR 3/2/1)"] = property_type_str
        print(f"🏠 Property type: {property_type_str}")

        # --- ENHANCED Agent Information Extraction ---
        try:
            agent_name = None
            agent_email = None

            soup = BeautifulSoup(driver.page_source, "html.parser")
            page_source = driver.page_source

            print("🔍 Starting enhanced agent contact extraction...")

            # Method 1: Look for agent information in structured data/JSON
            try:
                # Common JSON patterns for agent data
                agent_json_patterns = [
                    r'"agentName"\s*:\s*"([^"]+)"',
                    r'"listingAgentName"\s*:\s*"([^"]+)"',
                    r'"primaryAgent"\s*{\s*"name"\s*:\s*"([^"]+)"',
                    r'"agent"\s*:\s*{\s*"name"\s*:\s*"([^"]+)"',
                    r'"displayName"\s*:\s*"([^"]+)".*?"agentLicenseNumber"',
                    r'"fullName"\s*:\s*"([^"]+)".*?"isAgent"\s*:\s*true'
                ]

                for pattern in agent_json_patterns:
                    match = re.search(pattern, page_source, re.DOTALL)
                    if match:
                        potential_name = match.group(1).strip()
                        # Validate name (should be reasonable length and contain letters)
                        if 2 < len(potential_name) < 50 and re.search(r'[a-zA-Z]', potential_name):
                            agent_name = potential_name
                            print(f"👤 Agent name (from JSON): {agent_name}")
                            break

                # Look for email in JSON
                if agent_name:
                    # Look for email associated with the agent
                    email_patterns = [
                        rf'"{re.escape(agent_name)}".*?"email"\s*:\s*"([^"]+@[^"]+)"',
                        r'"email"\s*:\s*"([^"]+@[^"]+)".*?"isAgent"\s*:\s*true',
                        r'"agentEmail"\s*:\s*"([^"]+@[^"]+)"'
                    ]

                    for pattern in email_patterns:
                        match = re.search(pattern, page_source, re.DOTALL | re.IGNORECASE)
                        if match:
                            potential_email = match.group(1).strip()
                            if '@' in potential_email and '.' in potential_email:
                                agent_email = potential_email
                                print(f"📧 Agent email (from JSON): {agent_email}")
                                break

            except Exception as e:
                print(f"⚠️ Error extracting agent from JSON: {e}")

            # Method 2: Look for agent contact info in HTML elements
            if not agent_name or not agent_email:
                try:
                    # Look for agent sections/cards
                    agent_selectors = [
                        "[data-rf-test-id*='agent']",
                        ".agent-card", ".agent-info", ".agent-details",
                        "[class*='Agent']", "[class*='agent']",
                        ".listing-agent", ".contact-agent"
                    ]

                    for selector in agent_selectors:
                        try:
                            agent_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            for element in agent_elements:
                                element_text = element.text.strip()
                                if len(element_text) < 10:  # Skip very short elements
                                    continue

                                # Look for name in element text
                                if not agent_name:
                                    # Try to extract name from common patterns
                                    name_patterns = [
                                        r'Listed by\s+([^\n\r\•]+)',
                                        r'Agent:\s*([^\n\r\•]+)',
                                        r'Contact\s+([^\n\r\•]+)',
                                        r'^([A-Z][a-z]+\s+[A-Z][a-z]+)',  # FirstName LastName at start
                                    ]

                                    for pattern in name_patterns:
                                        match = re.search(pattern, element_text, re.MULTILINE)
                                        if match:
                                            potential_name = match.group(1).strip()
                                            # Clean up common suffixes/prefixes
                                            cleaned_name = re.sub(r'\s+(•|at|with|from).*$', '', potential_name,
                                                                  flags=re.IGNORECASE)
                                            if 2 < len(cleaned_name) < 50 and re.search(r'[a-zA-Z]', cleaned_name):
                                                agent_name = cleaned_name
                                                print(f"👤 Agent name (from HTML element): {agent_name}")
                                                break

                                # Look for email in element
                                if not agent_email:
                                    email_matches = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
                                                               element_text)
                                    if email_matches:
                                        # Take the first valid-looking email
                                        for email in email_matches:
                                            if not email.endswith('.jpg') and not email.endswith(
                                                    '.png'):  # Skip image file references
                                                agent_email = email
                                                print(f"📧 Agent email (from HTML element): {agent_email}")
                                                break

                                if agent_name and agent_email:
                                    break

                            if agent_name and agent_email:
                                break
                        except Exception as elem_e:
                            print(f"⚠️ Error processing agent element: {elem_e}")
                            continue

                except Exception as e:
                    print(f"⚠️ Error in HTML agent extraction: {e}")

            # Method 3: Search page text for email patterns
            if not agent_email:
                try:
                    page_text = soup.get_text()
                    # Find all emails in page
                    all_emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', page_text)

                    # Filter out common non-agent emails
                    filtered_emails = []
                    exclude_patterns = [
                        r'support@', r'info@', r'contact@', r'hello@', r'team@',
                        r'@redfin\.com$', r'@zillow\.com$', r'@realtor\.com$',
                        r'noreply', r'donotreply', r'admin@'
                    ]

                    for email in all_emails:
                        is_excluded = False
                        for pattern in exclude_patterns:
                            if re.search(pattern, email, re.IGNORECASE):
                                is_excluded = True
                                break
                        if not is_excluded and len(email) < 50:  # Reasonable length
                            filtered_emails.append(email)

                    if filtered_emails:
                        agent_email = filtered_emails[0]  # Take the first reasonable email
                        print(f"📧 Agent email (from page text): {agent_email}")

                except Exception as e:
                    print(f"⚠️ Error extracting email from page text: {e}")

            # Method 4: Fallback name extraction from page text
            if not agent_name:
                try:
                    page_text = soup.get_text()

                    # Look for "Listed by" or similar patterns in full page text
                    name_patterns = [
                        r'Listed by\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                        r'Contact\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                        r'Agent:\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                        r'Listing Agent:\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)'
                    ]

                    for pattern in name_patterns:
                        matches = re.findall(pattern, page_text)
                        if matches:
                            potential_name = matches[0].strip()
                            if 2 < len(potential_name) < 50:
                                agent_name = potential_name
                                print(f"👤 Agent name (from page text fallback): {agent_name}")
                                break

                except Exception as e:
                    print(f"⚠️ Error in fallback name extraction: {e}")

            # Combine agent contact info
            contact_info_parts = []
            if agent_name:
                contact_info_parts.append(agent_name)
            if agent_email:
                contact_info_parts.append(agent_email)

            if contact_info_parts:
                # Format as requested: Name + Email on separate lines or joined
                if len(contact_info_parts) == 2:
                    agent_contact = f"{contact_info_parts[0]}\n{contact_info_parts[1]}"
                else:
                    agent_contact = contact_info_parts[0]

                data["seller/agent/wholesaler/MLS"] = agent_contact
                print(f"👥 Final agent contact info: {agent_contact}")
            else:
                data["seller/agent/wholesaler/MLS"] = "Check Listing"
                print("⚠️ No agent contact info found")

        except Exception as e:
            print(f"⚠️ Error extracting agent contact info: {e}")
            data["seller/agent/wholesaler/MLS"] = "Check Listing"

        # --- Lot Size ---
        try:
            # Try to extract lot size from JSON data
            lot_match = re.search(r'"lotSize"\s*:\s*(\d+)', driver.page_source)
            if lot_match:
                data["lot size"] = int(lot_match.group(1))
                print(f"🏞️ Lot Size (from JSON): {data['lot size']}")
            else:
                # Try to find lot size in HTML
                soup = BeautifulSoup(driver.page_source, "html.parser")

                # Look for lot size patterns
                lot_patterns = [
                    r"Lot Size\s*:?\s*([0-9,]+)\s*sq\s*ft",
                    r"Lot\s*:?\s*([0-9,]+)\s*sq\s*ft",
                    r"([0-9,]+)\s*sq\s*ft\s*lot",
                ]

                page_text = soup.get_text()
                for pattern in lot_patterns:
                    match = re.search(pattern, page_text, re.IGNORECASE)
                    if match:
                        lot_size_str = match.group(1).replace(",", "")
                        if lot_size_str.isdigit():
                            data["lot size"] = int(lot_size_str)
                            print(f"🏞️ Lot Size (from HTML): {data['lot size']}")
                            break
        except Exception as e:
            print(f"⚠️ Error extracting Lot Size: {e}")

        # --- Year Built ---
        try:
            # First try to find year built in JSON data
            year_match = re.search(r'"yearBuilt"\s*:\s*(\d{4})', driver.page_source)
            if year_match:
                data["year built"] = year_match.group(1)
                print(f"🏗 Year Built (from JSON): {data['year built']}")
            else:
                # Fallback to HTML parsing
                soup = BeautifulSoup(driver.page_source, "html.parser")

                # Try different methods to find year built
                spans = soup.find_all("span", string=re.compile(r"Year Built", re.IGNORECASE))
                for span in spans:
                    parent = span.find_parent()
                    if parent:
                        next_val = parent.find_next_sibling()
                        if next_val:
                            text = next_val.get_text(strip=True)
                            if text.isdigit() and len(text) == 4:
                                data["year built"] = int(text)
                                print(f"🏗 Year Built (from span): {text}")
                                break

                if "year built" not in data:
                    lis = soup.find_all("li")
                    for li in lis:
                        if "Year Built" in li.text:
                            match = re.search(r"(\d{4})", li.text)
                            if match:
                                data["year built"] = match.group(1)
                                print(f"🏗 Year Built (from li): {data['year built']}")
                                break
        except Exception as e:
            print(f"⚠️ Error extracting Year Built: {e}")

        return data

    except Exception as e:
        print(f"❌ Failed to scrape Redfin: {e}")
        return {}
    finally:
        try:
            driver.quit()
        except:
            pass


def search_zillow_url(address):
    """Return the first Zillow property URL found for `address` (no API key)."""
    print(f"🔍 Searching Zillow listing: {address}")
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
    )

    try:
        service = ChromeService()
        driver = webdriver.Chrome(service=service, options=options)

        queries = [
            f"{address} site:zillow.com",
            f"\"{address}\" site:zillow.com",
            f"{address.replace(',', '')} zillow",
        ]
        for q in queries:
            driver.get(f"https://duckduckgo.com/?q={quote_plus(q)}")
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='zillow.com']"))
                )
                for a in driver.find_elements(By.CSS_SELECTOR, "a[href*='zillow.com']"):
                    href = a.get_attribute("href")
                    if href and "/homedetails/" in href:
                        print(f"✅ Zillow URL found: {href}")
                        return href.split("?")[0]  # strip tracking params
            except Exception:
                pass
        print("❌ Zillow link not found.")
        return None
    finally:
        try:
            driver.quit()
        except:
            pass


def get_zillow_data(url):
    """
    Enhanced Zillow scraper with better debugging and more extraction methods.
    """
    print(f"🌐 Scraping Zillow data: {url}")
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
    )

    data = {}
    try:
        service = ChromeService()
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(url)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # Add extra wait for dynamic content
        time.sleep(3)

        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        print("🔍 Starting Zillow data extraction...")

        # ── Enhanced helper functions ──────────────────────────
        def _extract_number(text):
            """Extract number from text like '$123,456' or '123456'"""
            if not text:
                return None
            # Remove all non-digit characters except decimal points
            cleaned = re.sub(r'[^\d.]', '', str(text))
            if cleaned and cleaned.replace('.', '').isdigit():
                return int(float(cleaned))
            return None

        def _record(key, val, label):
            if val and key not in data:
                if isinstance(val, str):
                    val = _extract_number(val)
                if val and isinstance(val, (int, float)):
                    data[key] = int(val)
                    print(f"{label} → ${data[key]:,}")
                    return True
            return False

        # ── Method 1: Enhanced JSON extraction ────────────────
        print("🔍 Trying Method 1: JSON data extraction...")
        try:
            # Look for multiple JSON script patterns
            json_scripts = soup.find_all("script", type="application/json")
            json_scripts.extend(soup.find_all("script", id=lambda x: x and "json" in x.lower()))
            json_scripts.extend(
                soup.find_all("script", string=lambda x: x and ("zestimate" in x.lower() or "rent" in x.lower())))

            for script in json_scripts:
                if script.string:
                    try:
                        json_data = json.loads(script.string)

                        # Deep search for zestimate values
                        def find_in_json(obj, target_keys):
                            results = {}
                            if isinstance(obj, dict):
                                for key, value in obj.items():
                                    key_lower = key.lower()
                                    if any(target in key_lower for target in target_keys):
                                        if isinstance(value, (int, str)) and str(value).replace(',', '').replace('$',
                                                                                                                 '').isdigit():
                                            results[key] = value
                                    elif isinstance(value, (dict, list)):
                                        results.update(find_in_json(value, target_keys))
                            elif isinstance(obj, list):
                                for item in obj:
                                    results.update(find_in_json(item, target_keys))
                            return results

                        # Search for zestimate patterns
                        zest_results = find_in_json(json_data, ['zestimate', 'estimated', 'value'])
                        rent_results = find_in_json(json_data, ['rent', 'rental'])

                        print(f"📊 Found potential zestimate values: {zest_results}")
                        print(f"🏠 Found potential rent values: {rent_results}")

                        # Try to assign values
                        for key, value in zest_results.items():
                            if 'rent' not in key.lower() and _extract_number(value) and _extract_number(value) > 10000:
                                _record("ARV estimated/appraised", value, "🏷️ Zestimate (JSON)")
                                break

                        for key, value in rent_results.items():
                            if 'zestimate' in key.lower() and _extract_number(value) and _extract_number(value) > 500:
                                _record("market rent", value, "💸 Rent Zestimate (JSON)")
                                break

                    except json.JSONDecodeError:
                        continue
        except Exception as e:
            print(f"⚠️ JSON extraction failed: {e}")

        # ── Method 2: Enhanced CSS selectors ──────────────────
        print("🔍 Trying Method 2: CSS selectors...")
        try:
            # Updated selectors for 2025 Zillow structure
            zestimate_selectors = [
                "[data-testid='zestimate-value']",
                "[data-testid='home-value']",
                "[data-testid='property-value']",
                ".zestimate-value",
                ".home-estimate-value",
                "[class*='Zestimate'] [class*='value']",
                "[aria-label*='Zestimate']",
                "span:contains('Zestimate')",
                ".price-summary .price",
                "[data-cy='zestimate-value']"
            ]

            for selector in zestimate_selectors:
                try:
                    elements = soup.select(selector)
                    for elem in elements:
                        text = elem.get_text(strip=True)
                        if _extract_number(text) and _extract_number(text) > 10000:
                            if _record("ARV estimated/appraised", text, f"🏷️ Zestimate ({selector})"):
                                break
                    if "ARV estimated/appraised" in data:
                        break
                except:
                    continue

            # Rent estimate selectors
            rent_selectors = [
                "[data-testid='rentZestimate-value']",
                "[data-testid='rent-estimate']",
                "[data-testid='rental-value']",
                ".rent-zestimate",
                ".rental-estimate",
                "[class*='RentZestimate'] [class*='value']",
                "[aria-label*='Rent Zestimate']",
                "span:contains('Rent Zestimate')",
                "[data-cy='rent-zestimate']"
            ]

            for selector in rent_selectors:
                try:
                    elements = soup.select(selector)
                    for elem in elements:
                        text = elem.get_text(strip=True)
                        if _extract_number(text) and 500 < _extract_number(text) < 10000:
                            if _record("market rent", text, f"💸 Rent Zestimate ({selector})"):
                                break
                    if "market rent" in data:
                        break
                except:
                    continue

        except Exception as e:
            print(f"⚠️ CSS selector method failed: {e}")

        # ── Method 3: Enhanced regex patterns ─────────────────
        print("🔍 Trying Method 3: Regex patterns...")
        try:
            # More comprehensive regex patterns for 2025
            zestimate_patterns = [
                r'Zestimate[®\s]*:?\s*\$?([\d,]+)',
                r'Home\s+value[:\s]*\$?([\d,]+)',
                r'Estimated\s+value[:\s]*\$?([\d,]+)',
                r'"zestimate"[:\s]*\$?([\d,]+)',
                r'Property\s+value[:\s]*\$?([\d,]+)',
                r'Current\s+estimate[:\s]*\$?([\d,]+)'
            ]

            for pattern in zestimate_patterns:
                matches = re.findall(pattern, html, re.IGNORECASE)
                for match in matches:
                    val = _extract_number(match)
                    if val and val > 10000:
                        if _record("ARV estimated/appraised", val, f"🏷️ Zestimate (regex: {pattern[:20]}...)"):
                            break
                if "ARV estimated/appraised" in data:
                    break

            rent_patterns = [
                r'Rent\s+Zestimate[®\s]*:?\s*\$?([\d,]+)',
                r'Rental\s+estimate[:\s]*\$?([\d,]+)',
                r'Monthly\s+rent[:\s]*\$?([\d,]+)',
                r'"rentZestimate"[:\s]*\$?([\d,]+)',
                r'Estimated\s+rent[:\s]*\$?([\d,]+)'
            ]

            for pattern in rent_patterns:
                matches = re.findall(pattern, html, re.IGNORECASE)
                for match in matches:
                    val = _extract_number(match)
                    if val and 500 < val < 10000:
                        if _record("market rent", val, f"💸 Rent Zestimate (regex: {pattern[:20]}...)"):
                            break
                if "market rent" in data:
                    break

        except Exception as e:
            print(f"⚠️ Regex extraction failed: {e}")

        # ── Method 4: Page text analysis ──────────────────────
        if not data:
            print("🔍 Trying Method 4: Full page text analysis...")
            try:
                page_text = soup.get_text()

                # Look for dollar amounts in reasonable ranges
                all_prices = re.findall(r'\$[\d,]+', page_text)
                print(f"💰 Found price candidates: {all_prices[:10]}...")  # Show first 10

                # Categorize by likely range
                potential_home_values = []
                potential_rents = []

                for price in all_prices:
                    val = _extract_number(price)
                    if val:
                        if 50000 <= val <= 2000000:  # Reasonable home value range
                            potential_home_values.append(val)
                        elif 500 <= val <= 10000:  # Reasonable rent range
                            potential_rents.append(val)

                # Take most common or median values
                if potential_home_values and "ARV estimated/appraised" not in data:
                    # Use the most common value or median
                    from collections import Counter
                    if len(potential_home_values) > 1:
                        most_common = Counter(potential_home_values).most_common(1)[0][0]
                        _record("ARV estimated/appraised", most_common, "🏷️ Zestimate (text analysis)")
                    else:
                        _record("ARV estimated/appraised", potential_home_values[0], "🏷️ Zestimate (text analysis)")

                if potential_rents and "market rent" not in data:
                    from collections import Counter
                    if len(potential_rents) > 1:
                        most_common = Counter(potential_rents).most_common(1)[0][0]
                        _record("market rent", most_common, "💸 Rent Zestimate (text analysis)")
                    else:
                        _record("market rent", potential_rents[0], "💸 Rent Zestimate (text analysis)")

            except Exception as e:
                print(f"⚠️ Text analysis failed: {e}")

        # ── Final results ──────────────────────────────────────
        if data:
            print(f"✅ Successfully extracted {len(data)} values from Zillow")
        else:
            print("❌ No data extracted from Zillow")
            # Save HTML for debugging
            with open("zillow_debug.html", "w", encoding="utf-8") as f:
                f.write(html)
                print("🐛 Saved HTML to zillow_debug.html for debugging")

        return data

    except Exception as e:
        print(f"⚠️ Zillow scrape failed: {e}")
        return {}
    finally:
        try:
            driver.quit()
        except:
            pass


def autofill_column(file_path, col_letter):
    print(f"📄 Opening workbook: {file_path}")
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active

    col_idx = column_index_from_string(col_letter)
    print(f"🧩 Targeting column '{col_letter}' (index {col_idx})")

    # Always grab the address from row 1 (needed for Zillow too)
    address_cell = ws.cell(row=1, column=col_idx)
    address = str(address_cell.value).strip() if address_cell.value else ""

    # Check if there's already a valid Redfin link in row 3
    link_cell = ws.cell(row=3, column=col_idx)
    existing_link = str(link_cell.value).strip() if link_cell.value else ""

    if is_valid_redfin_url(existing_link):
        print(f"✅ Found existing valid Redfin link: {existing_link}")
        link = existing_link
    else:
        print("🔍 No valid link found, searching for address...")
        if not address or address.lower() == "none":
            print("❌ No address found in row 1")
            return

        print(f"🏠 Found address: {address}")
        link = search_redfin_url(address)
        if not link:
            print("❌ Could not find Redfin listing for this address")
            return

        # Update the link cell only if it was empty
        if not existing_link:
            link_cell.value = link
            link_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            link_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            print(f"✅ Updated empty link cell with: {link}")

    print(f"🔗 Using link: {link}")

    if not is_valid_redfin_url(link):
        print(f"⚠️ Invalid or unsupported link: {link}")
        return

    # Get Redfin data
    print("🔍 Extracting Redfin data...")
    info = get_redfin_data(link)

    # ──────────────────────────────────────────────────────────
    # 🌟 ENHANCED: Fetch ARV & rent numbers from Zillow with better error handling
    # ──────────────────────────────────────────────────────────
    print("🔍 Searching for Zillow listing...")
    zillow_url = search_zillow_url(address)
    if zillow_url:
        print(f"✅ Found Zillow URL: {zillow_url}")
        print("🔍 Extracting Zillow data...")
        zillow_info = get_zillow_data(zillow_url)
        if zillow_info:
            print(f"✅ Zillow data extracted: {zillow_info}")
            # Merge Zillow data into main info
            info.update(zillow_info)
        else:
            print("⚠️ No data returned from Zillow")
    else:
        print("⚠️ Zillow link not found – ARV & rent will stay blank if labels exist.")

    if not info:
        print("⚠️ No data returned from either source.")
        return

    print(f"🎯 Total data available: {info}")

    # First, let's see what labels exist in the Excel file
    excel_labels = []
    print("📋 Reading Excel labels...")
    for row in range(4, 50):
        label_cell = ws.cell(row=row, column=1)
        if label_cell.value:
            label = str(label_cell.value).strip()
            excel_labels.append((row, label))

    fields_found = 0

    # FIXED: Only match specific fields we want to fill, let everything else be copied from column B
    for row, excel_label in excel_labels:
        excel_label_lower = excel_label.lower()

        # Only fill these specific fields - everything else should come from column B
        matched_value = None
        matched_key = None

        # Method 1: Exact match
        if excel_label in info:
            matched_value = info[excel_label]
            matched_key = excel_label
        # Method 2: Case insensitive match
        else:
            for k, v in info.items():
                if k.lower() == excel_label_lower:
                    matched_value = v
                    matched_key = k
                    break

        if matched_value is not None:
            cell = ws.cell(row=row, column=col_idx)

            # Skip if the sheet already has a value
            existing_val = str(cell.value).strip() if cell.value else ""
            if existing_val:
                print(f"⏭️ Row {row}: '{excel_label}' already has value: {existing_val}")
                continue

            if isinstance(matched_value, str) and matched_value.isdigit():
                matched_value = int(matched_value)

            cell.number_format = "General"
            cell.value = matched_value

            # Styling
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Special formatting for prices
            if "price" in excel_label.lower() or "arv" in excel_label.lower() or "rent" in excel_label.lower():
                if isinstance(matched_value, (int, float)) and matched_value > 0:
                    cell.number_format = '"$"#,##0'

            print(f"✏️ Row {row}: '{excel_label}' = {matched_value} (from {matched_key})")
            fields_found += 1

    if fields_found == 0:
        print("⚠️ No matching labels found.")
        print("📋 Available data keys:", list(info.keys()))
        print("📋 Excel labels found:", [label for _, label in excel_labels])
    else:
        print(f"✅ Filled {fields_found} fields.")

    # Continue with the rest of the function (copying from column B, saving, etc.)
    print(f"🔄 Copying empty cells from column B to column {col_letter} including formulas and fill colors...")
    for row in range(4, 100):
        source_cell = ws.cell(row=row, column=2)  # Column B
        target_cell = ws.cell(row=row, column=col_idx)

        if source_cell.value is not None and (target_cell.value is None or str(target_cell.value).strip() == ""):
            # Handle formula copy with column letter adaptation
            if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                formula = source_cell.value
                source_col_letter = "B"
                target_col_letter = col_letter.upper()

                # Replace only whole cell references (e.g. B9 → C9), not parts of strings like "BATH"
                adjusted_formula = re.sub(
                    rf'\b{source_col_letter}(\d+)\b',
                    rf'{target_col_letter}\1',
                    formula
                )
                target_cell.value = adjusted_formula
            else:
                target_cell.value = source_cell.value

            target_cell.number_format = source_cell.number_format
            target_cell.fill = copy(source_cell.fill)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.border = copy(source_cell.border)

    print("✅ Finished copying values and formatting from column B.")

    try:
        print("🔍 Attempting to fetch real comparables...")
        comps = get_redfin_comps(address, radius_miles=1, sold_within_days=365)

        if not comps:
            print("❌ No real comps found")
        else:
            log_comp_buckets(address, comps)
    except Exception as e:
        print(f"❌ Comp fetch failed: {e}")

    try:
        wb.save(file_path)
        print("✅ File saved successfully.")
    except Exception as e:
        print(f"❌ Failed to save file: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python autofill.py <COLUMN_LETTER> <EXCEL_PATH>")
        sys.exit(1)

    col_letter = sys.argv[1]
    file_path = sys.argv[2]
    print(f"🧩 Column: {col_letter}")
    print(f"📄 File:   {file_path}")

    autofill_column(file_path, col_letter)
    print("🏁 Done.")

